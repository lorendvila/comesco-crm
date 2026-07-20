#!/usr/bin/env python3
"""
Importador de tarifas COMESCO desde el Excel maestro.

Fuente de verdad: data/MAESTRO DE PRODUCTOS _ COMESCO.xlsx
  - 3 pestañas (RETAIL / FOOD SERVICE / INDUSTRIA), mismas filas por producto.
  - Emparejamiento por columna SKU (col M), que coincide con referencias.sku del CRM.
  - Columna G = coste hasta almacén con IVA -> referencias.coste_almacen_cop
  - Columna J = precio unitario SIN IVA por canal -> precio_<canal>_cop
      (RETAIL lleva el PAC 10% incluido en J; FS/INDUSTRIA no.)

Uso:
  python scripts/import_tarifas.py            # imprime el SQL por stdout
  python scripts/import_tarifas.py > out.sql  # lo vuelca a fichero

No escribe en Supabase directamente (referencias tiene RLS de solo-admin);
genera el SQL y se aplica con permisos. Al cambiar el Excel, relanzar y reaplicar.
"""
import openpyxl, sys, os

PATH = os.path.join(os.path.dirname(__file__), "..", "data", "MAESTRO DE PRODUCTOS _ COMESCO.xlsx")
CANAL_COL = {
    "RETAIL": "precio_retail_cop",
    "FOOD SERVICE": "precio_food_service_cop",
    "INDUSTRIA": "precio_industria_cop",
}
COL_SKU, COL_G, COL_J = 13, 7, 10  # M, G, J

def main():
    wb = openpyxl.load_workbook(PATH, data_only=True)
    datos = {}  # sku -> {"coste": G, "precio_retail_cop": .., ...}
    for tab, col in CANAL_COL.items():
        if tab not in wb.sheetnames:
            sys.exit(f"Falta la pestaña {tab}")
        ws = wb[tab]
        for r in range(2, ws.max_row + 1):
            sku = ws.cell(r, COL_SKU).value
            if not sku:
                continue
            sku = str(sku).strip()
            g = ws.cell(r, COL_G).value
            j = ws.cell(r, COL_J).value
            if j is None:
                continue
            d = datos.setdefault(sku, {"coste": None})
            d[col] = round(float(j))
            if g is not None:
                d["coste"] = round(float(g), 2)

    if not datos:
        sys.exit("No se leyó ninguna fila con SKU.")

    print("-- Generado por scripts/import_tarifas.py desde el Excel maestro")
    print(f"-- {len(datos)} referencias\n")
    for sku in sorted(datos):
        d = datos[sku]
        sets = [
            f"precio_retail_cop = {d.get('precio_retail_cop', 'NULL')}",
            f"precio_food_service_cop = {d.get('precio_food_service_cop', 'NULL')}",
            f"precio_industria_cop = {d.get('precio_industria_cop', 'NULL')}",
        ]
        if d["coste"] is not None:
            sets.append(f"coste_almacen_cop = {d['coste']}")
        print(f"UPDATE referencias SET {', '.join(sets)} WHERE sku = '{sku}';")

if __name__ == "__main__":
    main()
